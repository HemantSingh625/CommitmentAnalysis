import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import json

try:
    with open('config.json', 'r') as f:
        config = json.load(f)
        PLANT_CODES = config.get('plant_codes', ['742'])
        TARGET_PRODUCT = config.get('product_category', 'CRCA')
except Exception:
    PLANT_CODES = ['742']
    TARGET_PRODUCT = 'CRCA'

plant_str = ", ".join(PLANT_CODES)
TITLE_SUFFIX = f"({TARGET_PRODUCT}, Plant {plant_str})"

MONTHS = ['April', 'May', 'June', 'July', 'August', 'September', 'October',
          'November', 'December', 'January', 'February', 'March']

COMMIT_FILE = "Cleaned_Commitments_Data.xlsx"
PROD_FILE = "Cleaned_Production_Data.xlsx"
OUTPUT_FILE = "Master_Compliance_Report.xlsx"

def build_master_report():
    print(f"Reading {COMMIT_FILE} and {PROD_FILE}...")
    
    try:
        commit_xls = pd.ExcelFile(COMMIT_FILE)
        prod_xls = pd.ExcelFile(PROD_FILE)
    except Exception as e:
        raise Exception(f"Phase 3 Error: Could not read intermediate files. {e}")

    c_sheets = [s for s in commit_xls.sheet_names if s in MONTHS]
    p_sheets = [s for s in prod_xls.sheet_names if s in MONTHS]

    if not c_sheets:
        raise Exception("Phase 3 Error: The cleaned commitment data contains no valid month sheets. This means none of your uploaded commitment files matched the selected Profile (e.g., wrong P Codes or Plants).")
    if not p_sheets:
        raise Exception("Phase 3 Error: The cleaned production data contains no valid month sheets. This means your Production file had 0 rows matching the selected Profile (Check if you selected '743&744' but uploaded the '742' file).")

    print("Concatenating monthly data...")
    commit_dfs = []
    for s in c_sheets:
        df = pd.read_excel(commit_xls, sheet_name=s)
        if 'Month' not in df.columns:
            df.insert(0, 'Month', s)
        commit_dfs.append(df)
        
    prod_dfs = []
    for s in p_sheets:
        df = pd.read_excel(prod_xls, sheet_name=s)
        prod_dfs.append(df)

    master_df = pd.concat(commit_dfs, ignore_index=True)
    prod_df = pd.concat(prod_dfs, ignore_index=True)

    print("Centralizing and standardizing customer names across both datasets...")
    import difflib
    import re
    
    def standardize_names_unified(commit_series, prod_series, threshold=0.90):
        c_names = commit_series.dropna().unique().tolist() if commit_series is not None else []
        p_names = prod_series.dropna().unique().tolist() if prod_series is not None else []
        all_unique = list(set(c_names + p_names))
        
        standard_map = {}
        
        def normalize(n):
            n = str(n).upper()
            n = re.sub(r'\b(LTD|LIMITED|LIMITE|LIMIT|PVT|PRIVATE|CO|COMPANY|AND|LIMI|CORP|CORPORATION|INC|PROCESSING|PROCESSORS?|MANUFACTURING|MANUFACTU|MANUFACTURE)\b', '', n)
            n = re.sub(r'[&.,\-\s()]', '', n)
            return n
            
        for name in all_unique:
            if not str(name).strip():
                continue
            norm_name = normalize(name)
            match_found = False
            for std_name in set(standard_map.values()):
                norm_std = normalize(std_name)
                if difflib.SequenceMatcher(None, norm_name, norm_std).ratio() >= threshold:
                    standard_map[name] = std_name
                    match_found = True
                    break
            if not match_found:
                standard_map[name] = str(name).strip().upper()
                
        return standard_map

    # Find the Customer Name columns dynamically in both DataFrames
    c_cust_col = next((c for c in master_df.columns if 'customer' in str(c).lower() or 'cust' in str(c).lower()), None)
    p_cust_col = next((c for c in prod_df.columns if 'customer' in str(c).lower() or 'cust' in str(c).lower()), None)

    if c_cust_col and p_cust_col:
        unified_map = standardize_names_unified(master_df[c_cust_col], prod_df[p_cust_col])
        master_df['Clean Customer Name'] = master_df[c_cust_col].map(unified_map).fillna("UNKNOWN")
        prod_df['Clean Customer Name'] = prod_df[p_cust_col].map(unified_map).fillna("UNKNOWN")
    else:
        print("Warning: Could not find Customer Name columns to standardize.")

    def standardize_sales_org(org):
        org = str(org).strip().upper()
        if 'AUTO' in org or '3000' in org:
            return '3000'
        elif 'TUBE' in org or '4000' in org:
            return '4000'
        elif 'BPR' in org or '2000' in org:
            return '2000'
        elif 'IPP' in org or '1000' in org:
            return '1000'
        elif org == 'FREE STOCKS' or org == 'NAN' or org == '':
            return 'free stocks'
        return str(org).strip()

    c_sales_org = next((c for c in master_df.columns if 'sale' in str(c).lower() and 'org' in str(c).lower()), None)
    p_sales_org = next((c for c in prod_df.columns if 'sale' in str(c).lower() and 'org' in str(c).lower()), None)

    if c_sales_org:
        master_df[c_sales_org] = master_df[c_sales_org].apply(standardize_sales_org)
    if p_sales_org:
        prod_df[p_sales_org] = prod_df[p_sales_org].apply(standardize_sales_org)

    print(f"Writing base data to {OUTPUT_FILE}...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_master = wb.create_sheet("Master Data - All Months")
    for r_idx, row in enumerate(dataframe_to_rows(master_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_master.cell(row=r_idx, column=c_idx, value=value)
    
    ws_prod = wb.create_sheet("Production Data")
    for r_idx, row in enumerate(dataframe_to_rows(prod_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_prod.cell(row=r_idx, column=c_idx, value=value)

    m_headers = [c.value for c in ws_master[1]]
    p_headers = [c.value for c in ws_prod[1]]

    def find_col_letter(headers, keywords):
        for i in range(len(headers) - 1, -1, -1):
            h = headers[i]
            if h and all(k in str(h).lower() for k in keywords):
                return get_column_letter(i + 1)
        return None

    c_map = {
        'gr': find_col_letter(m_headers, ['planned gr target']),
        'cust': find_col_letter(m_headers, ['clean customer name']),
        'combo': find_col_letter(m_headers, ['so-item combo']),
        'pcode': find_col_letter(m_headers, ['p code']),
        'vert': find_col_letter(m_headers, ['vertical']),
        'thk': find_col_letter(m_headers, ['thk']),
        'width': find_col_letter(m_headers, ['width']),
        'plant': find_col_letter(m_headers, ['plant']),
        'sales_org': find_col_letter(m_headers, ['sale']),
        'month': find_col_letter(m_headers, ['month'])
    }
    
    p_map = {
        'qty': find_col_letter(p_headers, ['quantity']),
        'cust': find_col_letter(p_headers, ['clean customer name']),
        'combo': find_col_letter(p_headers, ['so-item combo']),
        'pcode': find_col_letter(p_headers, ['p code clean']),
        'thk': find_col_letter(p_headers, ['thick']),
        'width': find_col_letter(p_headers, ['width']),
        'plant': find_col_letter(p_headers, ['plant clean']),
        'sales_org': find_col_letter(p_headers, ['sale']),
        'month': find_col_letter(p_headers, ['month name'])
    }

    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")

    print("Adding dynamic analysis sheets...")
    
    ws1 = wb.create_sheet("Overall Compliance")
    ws1["A1"] = f"Overall FY Compliance {TITLE_SUFFIX}"
    ws1["A1"].font = title_font

    headers = ["Month", "Commitment (Planned GR)", "Production (Qty)", "Compliance %"]
    for i, h in enumerate(headers, start=1):
        c = ws1.cell(3, i, h)
        c.font = bold
        c.fill = header_fill

    c_gr = c_map.get('gr')
    p_qty = p_map.get('qty')
    
    c_month_col = c_map.get('month')
    p_month_col = p_map.get('month')

    r = 4
    for m in MONTHS:
        ws1.cell(r, 1, m)
        ws1.cell(r, 2, f"=SUMIF('Master Data - All Months'!{c_month_col}:{c_month_col},\"{m}\",'Master Data - All Months'!{c_gr}:{c_gr})")
        ws1.cell(r, 3, f"=SUMIF('Production Data'!{p_month_col}:{p_month_col},\"{m}\",'Production Data'!{p_qty}:{p_qty})")
        ws1.cell(r, 4, f"=IFERROR(C{r}/B{r}*100,0)")
        ws1.cell(r, 2).number_format = "#,##0.00"
        ws1.cell(r, 3).number_format = "#,##0.00"
        ws1.cell(r, 4).number_format = "0.0"
        r += 1

    last_row = r - 1
    fy_row = r + 1
    
    ws1.cell(fy_row, 1, "FY Total").font = bold
    ws1.cell(fy_row, 2, f"=SUM(B4:B{last_row})")
    ws1.cell(fy_row, 3, f"=SUM(C4:C{last_row})")
    ws1.cell(fy_row, 4, f"=IFERROR(C{fy_row}/B{fy_row}*100,0)")
    ws1.cell(fy_row, 4).number_format = "0.0"
    ws1.cell(fy_row, 2).number_format = "#,##0.00"
    ws1.cell(fy_row, 3).number_format = "#,##0.00"

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 15

    def build_analysis_sheet(sheet_name, row_field_c, row_field_p, title, width=25):
        if not c_map.get(row_field_c) or not p_map.get(row_field_p):
            print(f"Skipping sheet {sheet_name} due to missing columns.")
            return
        
        c_col = c_map[row_field_c]
        p_col = p_map[row_field_p]
        c_gr = c_map['gr']
        p_qty = p_map['qty']
        
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = f"{title} Monthly Breakdown {TITLE_SUFFIX}"
        ws["A1"].font = title_font
        
        headers = [title]
        for m in MONTHS:
            headers.extend([f"{m} Comm", f"{m} Prod", f"{m} %"])
        headers.extend(["FY Total Comm", "FY Total Prod", "FY Total %"])
        
        for i, h in enumerate(headers, start=1):
            c = ws.cell(3, i, h)
            c.font = bold
            c.fill = header_fill
            
        unique_keys = pd.concat([master_df.iloc[:, openpyxl.utils.column_index_from_string(c_col)-1], 
                               prod_df.iloc[:, openpyxl.utils.column_index_from_string(p_col)-1]]).dropna().astype(str).str.replace(r'\.0$', '', regex=True).unique()
                               
        r = 4
        for k in sorted(unique_keys):
            display_k = 'free stocks' if str(k).strip() == '' else k
            ws.cell(r, 1, display_k)
            col_idx = 2
            
            fy_comm_sum = []
            fy_prod_sum = []
            
            for m in MONTHS:
                # Comm
                c_cell = ws.cell(r, col_idx, f"=SUMIFS('Master Data - All Months'!{c_gr}:{c_gr},'Master Data - All Months'!{c_col}:{c_col},\"{k}\",'Master Data - All Months'!{c_month_col}:{c_month_col},\"{m}\")")
                c_cell.number_format = "#,##0.00"
                fy_comm_sum.append(get_column_letter(col_idx) + str(r))
                col_idx += 1
                
                # Prod
                p_cell = ws.cell(r, col_idx, f"=SUMIFS('Production Data'!{p_qty}:{p_qty},'Production Data'!{p_col}:{p_col},\"{k}\",'Production Data'!{p_month_col}:{p_month_col},\"{m}\")")
                p_cell.number_format = "#,##0.00"
                fy_prod_sum.append(get_column_letter(col_idx) + str(r))
                col_idx += 1
                
                # %
                pct_cell = ws.cell(r, col_idx, f"=IFERROR({get_column_letter(col_idx-1)}{r}/{get_column_letter(col_idx-2)}{r}*100,0)")
                pct_cell.number_format = "0.0"
                col_idx += 1
                
            # FY Totals
            fy_c_cell = ws.cell(r, col_idx, f"={'+'.join(fy_comm_sum) if fy_comm_sum else '0'}")
            fy_c_cell.number_format = "#,##0.00"
            col_idx += 1
            
            fy_p_cell = ws.cell(r, col_idx, f"={'+'.join(fy_prod_sum) if fy_prod_sum else '0'}")
            fy_p_cell.number_format = "#,##0.00"
            col_idx += 1
            
            fy_pct_cell = ws.cell(r, col_idx, f"=IFERROR({get_column_letter(col_idx-1)}{r}/{get_column_letter(col_idx-2)}{r}*100,0)")
            fy_pct_cell.number_format = "0.0"
            
            r += 1
            
        ws.column_dimensions["A"].width = width
        for i in range(2, len(headers) + 1):
            if (i - 1) % 3 == 0:
                ws.column_dimensions[get_column_letter(i)].width = 10  # % column
            else:
                ws.column_dimensions[get_column_letter(i)].width = 15  # Comm/Prod columns

    build_analysis_sheet("Product Compliance", 'pcode', 'pcode', "P Code", 20)
    build_analysis_sheet("Customer Compliance", 'cust', 'cust', "Customer Name", 35)
    build_analysis_sheet("Order Item Compliance", 'combo', 'combo', "SO-Item Combo", 25)
    build_analysis_sheet("Dimensional Compliance", 'thk', 'thk', "Thickness", 20)
    build_analysis_sheet("Width Compliance", 'width', 'width', "Width", 20)
    build_analysis_sheet("Plant Compliance", 'plant', 'plant', "Plant", 20)
    build_analysis_sheet("Sales Org Compliance", 'sales_org', 'sales_org', "Sales Organization", 25)

    print(f"Writing final {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print(f"Phase 3 Complete! {OUTPUT_FILE} created.")

if __name__ == "__main__":
    build_master_report()
