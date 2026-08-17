import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import json

try:
    with open('config.json', 'r') as f:
        config = json.load(f)
        PLANT_CODES = config.get('plant_codes', ['742'])
        TARGET_PRODUCT = config.get('product_category', 'CRCA')
except Exception:
    PLANT_CODES = ['742']
    TARGET_PRODUCT = 'CRCA'

plant_str = ", ".join(PLANT_CODES)
TITLE_SUFFIX = f"({TARGET_PRODUCT}, Plant {plant_str})"

MONTHS = ['April', 'May', 'June', 'July', 'August', 'September', 'October',
          'November', 'December', 'January', 'February', 'March']

COMMIT_FILE = "Cleaned_Commitments_Data.xlsx"
PROD_FILE = "Cleaned_Production_Data.xlsx"
OUTPUT_FILE = "Master_Compliance_Report.xlsx"

def build_master_report():
    print(f"Reading {COMMIT_FILE} and {PROD_FILE}...")
    
    try:
        commit_xls = pd.ExcelFile(COMMIT_FILE)
        prod_xls = pd.ExcelFile(PROD_FILE)
    except Exception as e:
        print(f"Error reading files: {e}")
        return

    c_sheets = [s for s in commit_xls.sheet_names if s in MONTHS]
    p_sheets = [s for s in prod_xls.sheet_names if s in MONTHS]

    if not c_sheets or not p_sheets:
        print("Error: Could not find valid month sheets.")
        return

    print("Concatenating monthly data...")
    commit_dfs = []
    for s in c_sheets:
        df = pd.read_excel(commit_xls, sheet_name=s)
        if 'Month' not in df.columns:
            df.insert(0, 'Month', s)
        commit_dfs.append(df)
        
    prod_dfs = []
    for s in p_sheets:
        df = pd.read_excel(prod_xls, sheet_name=s)
        prod_dfs.append(df)

    master_df = pd.concat(commit_dfs, ignore_index=True)
    prod_df = pd.concat(prod_dfs, ignore_index=True)

    print(f"Writing base data to {OUTPUT_FILE}...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_master = wb.create_sheet("Master Data - All Months")
    for r_idx, row in enumerate(dataframe_to_rows(master_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_master.cell(row=r_idx, column=c_idx, value=value)
    
    ws_prod = wb.create_sheet("Production Data")
    for r_idx, row in enumerate(dataframe_to_rows(prod_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_prod.cell(row=r_idx, column=c_idx, value=value)

    m_headers = [c.value for c in ws_master[1]]
    p_headers = [c.value for c in ws_prod[1]]

    def find_col_letter(headers, keywords):
        for i, h in enumerate(headers):
            if h and all(k in str(h).lower() for k in keywords):
                return get_column_letter(i + 1)
        return None

    c_map = {
        'gr': find_col_letter(m_headers, ['planned gr target']),
        'cust': find_col_letter(m_headers, ['clean customer name']),
        'combo': find_col_letter(m_headers, ['so-item combo']),
        'pcode': find_col_letter(m_headers, ['p code']),
        'vert': find_col_letter(m_headers, ['vertical']),
        'thk': find_col_letter(m_headers, ['thk'])
    }
    
    p_map = {
        'qty': find_col_letter(p_headers, ['quantity']),
        'cust': find_col_letter(p_headers, ['clean customer name']),
        'combo': find_col_letter(p_headers, ['so-item combo']),
        'pcode': find_col_letter(p_headers, ['p code clean']),
        'thk': find_col_letter(p_headers, ['thick'])
    }

    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")

    print("Adding dynamic analysis sheets...")
    
    ws1 = wb.create_sheet("Overall Compliance")
    ws1["A1"] = f"Overall FY Compliance {TITLE_SUFFIX}"
    ws1["A1"].font = title_font

    headers = ["Month", "Commitment (Planned GR)", "Production (Qty)", "Compliance %"]
    for i, h in enumerate(headers, start=1):
        c = ws1.cell(3, i, h)
        c.font = bold
        c.fill = header_fill

    c_gr = c_map.get('gr')
    p_qty = p_map.get('qty')
    
    c_month_col = 'A'
    p_month_col = 'AH'

    r = 4
    for m in MONTHS:
        ws1.cell(r, 1, m)
        ws1.cell(r, 2, f"=SUMIF('Master Data - All Months'!{c_month_col}:{c_month_col},\"{m}\",'Master Data - All Months'!{c_gr}:{c_gr})")
        ws1.cell(r, 3, f"=SUMIF('Production Data'!{p_month_col}:{p_month_col},\"{m}\",'Production Data'!{p_qty}:{p_qty})")
        ws1.cell(r, 4, f"=IFERROR(C{r}/B{r}*100,0)")
        ws1.cell(r, 2).number_format = "#,##0.00"
        ws1.cell(r, 3).number_format = "#,##0.00"
        ws1.cell(r, 4).number_format = "0.0"
        r += 1

    last_row = r - 1
    fy_row = r + 1
    
    ws1.cell(fy_row, 1, "FY Total").font = bold
    ws1.cell(fy_row, 2, f"=SUM(B4:B{last_row})")
    ws1.cell(fy_row, 3, f"=SUM(C4:C{last_row})")
    ws1.cell(fy_row, 4, f"=IFERROR(C{fy_row}/B{fy_row}*100,0)")
    ws1.cell(fy_row, 4).number_format = "0.0"
    ws1.cell(fy_row, 2).number_format = "#,##0.00"
    ws1.cell(fy_row, 3).number_format = "#,##0.00"

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 15

    def build_analysis_sheet(sheet_name, row_field_c, row_field_p, title, width=25):
        if not c_map.get(row_field_c) or not p_map.get(row_field_p):
            return
        
        c_col = c_map[row_field_c]
        p_col = p_map[row_field_p]
        c_gr = c_map['gr']
        p_qty = p_map['qty']
        
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = f"{title} {TITLE_SUFFIX}"
        ws["A1"].font = title_font
        
        for i, h in enumerate([title, "Commitment (Planned GR)", "Production (Qty)", "Compliance %"], start=1):
            c = ws.cell(3, i, h)
            c.font = bold
            c.fill = header_fill
            
        unique_keys = pd.concat([master_df.iloc[:, openpyxl.utils.column_index_from_string(c_col)-1], 
                               prod_df.iloc[:, openpyxl.utils.column_index_from_string(p_col)-1]]).dropna().unique()
                               
        r = 4
        for k in sorted(unique_keys):
            ws.cell(r, 1, k)
            ws.cell(r, 2, f"=SUMIF('Master Data - All Months'!{c_col}:{c_col},\"{k}\",'Master Data - All Months'!{c_gr}:{c_gr})")
            ws.cell(r, 3, f"=SUMIF('Production Data'!{p_col}:{p_col},\"{k}\",'Production Data'!{p_qty}:{p_qty})")
            ws.cell(r, 4, f"=IFERROR(C{r}/B{r}*100,0)")
            ws.cell(r, 2).number_format = "#,##0.00"
            ws.cell(r, 3).number_format = "#,##0.00"
            ws.cell(r, 4).number_format = "0.0"
            r += 1
            
        for col, w in zip("ABCD", [width, 25, 20, 15]):
            ws.column_dimensions[col].width = w

    build_analysis_sheet("Product Compliance", 'pcode', 'pcode', "P Code", 20)
    build_analysis_sheet("Customer Compliance", 'cust', 'cust', "Customer Name", 35)
    build_analysis_sheet("Order Item Compliance", 'combo', 'combo', "SO-Item Combo", 25)
    build_analysis_sheet("Dimensional Compliance", 'thk', 'thk', "Thickness", 20)

    print(f"Writing final {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print(f"Phase 3 Complete! {OUTPUT_FILE} created.")

if __name__ == "__main__":
    build_master_report()
